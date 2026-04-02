"""
Seed script: loads médicos data from 'Painel 2025.xlsx' and populates the database.

Reads:
  - NOME, CRM, ESPECIALIDADE, ENDEREÇO, BAIRRO
  - SEGUNDA-SEXTA columns with availability windows (e.g., "14x17" = 14:00-17:00)

Uses Nominatim (free) to resolve address → latitude, longitude.

Run with:
    cd backend && python seed_from_excel.py
"""
from __future__ import annotations

import sys
import os
import uuid
import datetime
import logging
import time
from typing import Optional

import openpyxl
from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut, GeocoderServiceError

# Allow imports from the backend directory
sys.path.insert(0, os.path.dirname(__file__))

from database import SessionLocal, Base, engine
import models  # noqa: F401 — registers all models with Base.metadata

from models.representante import Representante
from models.medico import Medico
from models.local_atendimento import LocalAtendimento
from models.disponibilidade import Disponibilidade


# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)


# Nominatim geocoder (rate limited to ~1 req/sec)
geolocator = Nominatim(user_agent="pharma_route_poc")


def is_orange_color(hex_color: str) -> bool:
    """
    Check if a hex color is orange-ish (RGB values suggest orange).
    Orange typically has high R, medium-high G, low B.
    Examples: FFA500, FFB366, FF9900, etc.
    """
    if not hex_color or len(hex_color) < 6:
        return False

    try:
        # Remove 'FF' prefix if present (alpha channel)
        color = hex_color.upper()
        if color.startswith("FF"):
            color = color[2:]

        if len(color) < 6:
            return False

        r = int(color[0:2], 16)
        g = int(color[2:4], 16)
        b = int(color[4:6], 16)

        # Orange: R > 200, G between 100-200, B < 100
        return r > 200 and 100 <= g <= 200 and b < 100
    except (ValueError, IndexError):
        return False


def get_cell_priority(cell) -> str:
    """
    Detect cell background color and return priority.
    Orange background = "A" (prioritário)
    Other/no color = "B"
    """
    try:
        if not cell or not cell.fill:
            return "B"

        fg_color = cell.fill.fgColor
        if not fg_color:
            return "B"

        # openpyxl represents colors as hex strings
        hex_color = getattr(fg_color, "rgb", None)
        if hex_color and is_orange_color(str(hex_color)):
            return "A"

    except Exception as e:
        logger.debug(f"Error checking cell color: {e}")

    return "B"


def parse_availability(availability_str: str) -> Optional[tuple[int, int]]:
    """
    Parse availability string like "14x17" or "14X17" into (hour_start, hour_end).
    Returns None if invalid or empty.
    """
    if not availability_str:
        return None

    availability_str = str(availability_str).strip().upper()
    if "X" not in availability_str:
        return None

    parts = availability_str.split("X")
    if len(parts) != 2:
        return None

    try:
        hour_start = int(parts[0].strip())
        hour_end = int(parts[1].strip())
        return (hour_start, hour_end)
    except ValueError:
        return None


def geocode_address(endereco: str, bairro: str) -> tuple[float, float] | None:
    """
    Resolve address + bairro to (latitude, longitude) using Nominatim.
    Returns None if unable to geocode.
    Includes rate limiting (~1 sec between requests).
    """
    if not endereco or not bairro:
        return None

    full_address = f"{endereco}, {bairro}, Rio de Janeiro, Brazil"

    try:
        # Rate limiting: sleep 1 sec between Nominatim requests
        time.sleep(1)
        location = geolocator.geocode(full_address, timeout=10)
        if location:
            return (location.latitude, location.longitude)
    except (GeocoderTimedOut, GeocoderServiceError) as e:
        logger.warning(f"Geocoder error for '{full_address}': {e}")
    except Exception as e:
        logger.warning(f"Unexpected geocoding error for '{full_address}': {e}")

    return None


def load_medicos_from_excel(filepath: str) -> list[dict]:
    """
    Load médicos data from Excel file.
    Returns list of dicts with keys: nome, crm, especialidade, endereco, bairro, disponibilidades, prioridade

    Prioridade is determined by SETOR column (col 1) color:
    - Orange background = "A" (prioritário)
    - Other/no color = "B"
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    medicos = []

    # Map day columns to day_of_week (0=Monday, 4=Friday)
    day_columns = {
        8: 0,   # SEGUNDA (col 9 in 1-indexed)
        9: 1,   # TERÇA
        10: 2,  # QUARTA
        11: 3,  # QUINTA
        12: 4,  # SEXTA
    }

    for row_idx in range(2, ws.max_row + 1):
        # Get cells (not just values) to access formatting
        row_cells = list(ws[row_idx])
        row_data = [cell.value for cell in row_cells]

        # Extract basic info (convert to str, handling None and float values)
        nome = str(row_data[2]) if row_data[2] else None  # Col 3
        crm = str(row_data[1]) if row_data[1] else None   # Col 2
        especialidade = str(row_data[3]) if row_data[3] else "Não especificado"  # Col 4
        endereco = str(row_data[4]) if row_data[4] else ""  # Col 5
        bairro = str(row_data[6]) if row_data[6] else ""  # Col 7

        # Skip if no name or CRM
        if not nome or not crm or nome == "None" or crm == "None":
            continue

        # Detect priority from SETOR column (col 1, index 0) color
        setor_cell = row_cells[0]  # Column A (index 0)
        prioridade = get_cell_priority(setor_cell)

        # Parse availability for each day
        disponibilidades = []
        for col_idx, day_of_week in day_columns.items():
            availability_str = row_data[col_idx]
            parsed = parse_availability(availability_str)

            if parsed:
                hour_start, hour_end = parsed
                disponibilidades.append({
                    "dia_semana": day_of_week,
                    "hora_inicio": datetime.time(hour_start, 0),
                    "hora_fim": datetime.time(hour_end, 0),
                })

        # Only include if has availability
        if not disponibilidades:
            continue

        medicos.append({
            "nome": nome.strip(),
            "crm": crm.strip(),
            "especialidade": especialidade.strip(),
            "endereco": endereco.strip(),
            "bairro": bairro.strip(),
            "prioridade": prioridade,
            "disponibilidades": disponibilidades,
        })

    logger.info(f"Loaded {len(medicos)} médicos from Excel")
    return medicos


def seed() -> None:
    """
    Main seed function: load from Excel and populate database.
    """
    Base.metadata.create_all(bind=engine)
    db = SessionLocal()

    try:
        # Check if already seeded
        from sqlalchemy import select
        existing = db.execute(select(Representante)).scalars().first()
        if existing:
            logger.warning("Database already seeded. Skipping.")
            return

        # Create default representante
        rep = Representante(
            id=str(uuid.uuid4()),
            nome="Representante Padrão",
            email="rep@pharma.com",
            regiao="Rio de Janeiro",
        )
        db.add(rep)
        db.flush()

        # Load médicos from Excel
        excel_path = os.path.join(os.path.dirname(__file__), "..", "Painel 2025.xlsx")
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        medicos_data = load_medicos_from_excel(excel_path)

        if not medicos_data:
            logger.warning("No médicos with availability found in Excel")
            return

        # Create medicos, locais, and disponibilidades
        medicos_created = 0
        medicos_priority_a = 0
        medicos_priority_b = 0
        locais_created = 0
        disps_created = 0
        medicos_skipped = 0

        for m_data in medicos_data:
            try:
                # Geocode address to get lat/lon
                lat_lon = geocode_address(m_data["endereco"], m_data["bairro"])

                if not lat_lon:
                    logger.warning(
                        f"Could not geocode {m_data['nome']} at "
                        f"{m_data['endereco']}, {m_data['bairro']} — skipping"
                    )
                    medicos_skipped += 1
                    continue

                latitude, longitude = lat_lon

                # Create médico
                medico = Medico(
                    id=str(uuid.uuid4()),
                    representante_id=rep.id,
                    nome=m_data["nome"],
                    crm=m_data["crm"],
                    especialidade=m_data["especialidade"],
                    prioridade=m_data["prioridade"],  # From Excel cell color (A=orange, B=other)
                )
                db.add(medico)
                db.flush()
                medicos_created += 1
                if m_data["prioridade"] == "A":
                    medicos_priority_a += 1
                else:
                    medicos_priority_b += 1

                # Create local de atendimento
                local = LocalAtendimento(
                    id=str(uuid.uuid4()),
                    medico_id=medico.id,
                    nome=f"Consultório/Clínica - {m_data['bairro']}",
                    endereco=f"{m_data['endereco']}, {m_data['bairro']}, Rio de Janeiro",
                    latitude=latitude,
                    longitude=longitude,
                    tipo="consultorio",  # Default type
                )
                db.add(local)
                db.flush()
                locais_created += 1

                # Create disponibilidades
                for disp_data in m_data["disponibilidades"]:
                    disp = Disponibilidade(
                        id=str(uuid.uuid4()),
                        local_id=local.id,
                        dia_semana=disp_data["dia_semana"],
                        hora_inicio=disp_data["hora_inicio"],
                        hora_fim=disp_data["hora_fim"],
                    )
                    db.add(disp)
                    disps_created += 1

                logger.info(
                    f"✓ {m_data['nome']} ({m_data['crm']}) "
                    f"at {latitude:.4f}, {longitude:.4f}"
                )

            except Exception as e:
                logger.error(f"Error processing médico {m_data['nome']}: {e}")
                medicos_skipped += 1
                continue

        db.commit()

        logger.info(
            f"\n=== SEEDING COMPLETE ===\n"
            f"Representante: 1\n"
            f"Médicos created: {medicos_created}\n"
            f"  ├─ Priority A (orange): {medicos_priority_a}\n"
            f"  └─ Priority B (other): {medicos_priority_b}\n"
            f"Médicos skipped (geocoding failed): {medicos_skipped}\n"
            f"Locais created: {locais_created}\n"
            f"Disponibilidades created: {disps_created}\n"
        )

    except Exception as e:
        db.rollback()
        logger.error(f"Fatal error during seeding: {e}")
        raise
    finally:
        db.close()


if __name__ == "__main__":
    seed()
